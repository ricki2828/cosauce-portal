"""Cashflow Excel parser service.

Parses the CoSauce cashflow workbook (ZAR-denominated) into structured data
ready for DB insertion.

Sheet structure (Cashflow sheet):
- Row 1: Header "COSAUCE - ZAR"
- Row 3: Column headers - Col 4 has "ZAR" (actuals), Cols 8-19 have datetime
          objects (2025-11 through 2026-10) for monthly forecast data
- Row 5: BANK BALANCE
- Row 6: NET CASH MOVEMENT
- Row 7: Operating Profit
- Row 8: Net Income
- Row 17: ZAR/NZD FX Rate
- Rows 31-42: Revenue by client
- Row 43: Total Revenue
- Rows 48-65: Expenses
- Row 66: Total Expenses
- Row 67: Operating Profit
- Row 72: Earnings Before Tax
- Row 76: Pre-Tax Income
- Rows 85-97: Loans section
- Rows 99-119: Vendors section
- Rows 121-127: Receivables section
- Row 131: Capex
- Rows 133-139: Owners section

Column mapping:
- Col 4 = October 2025 actuals (ZAR)
- Col 8 = Nov 2025 forecast
- Col 9 = Dec 2025 forecast
- ... Col 19 = Oct 2026 forecast
"""

from datetime import datetime
from typing import Optional
import uuid

import openpyxl


def _get_column_dates(ws) -> dict[int, tuple[int, int, bool]]:
    """
    Extract column-to-(year, month, is_actual) mapping from row 3.
    Col 4 is the actuals column (Oct 2025).
    Cols 8-19 are forecast months.
    Returns {col_index: (year, month, is_actual)}.
    """
    col_dates = {}

    # Column 4 is October 2025 actuals
    col_dates[4] = (2025, 10, True)

    # Columns 8-19 are forecast months
    for col in range(8, 20):
        cell = ws.cell(row=3, column=col).value
        if cell is not None:
            if isinstance(cell, datetime):
                col_dates[col] = (cell.year, cell.month, False)
            elif isinstance(cell, str):
                # Try to parse string date
                try:
                    dt = datetime.strptime(cell, "%Y-%m-%d")
                    col_dates[col] = (dt.year, dt.month, False)
                except ValueError:
                    pass

    return col_dates


def _read_row_values(ws, row: int, col_dates: dict) -> list[dict]:
    """Read a row's values across all date columns, returning list of entries."""
    entries = []
    for col, (year, month, is_actual) in col_dates.items():
        val = ws.cell(row=row, column=col).value
        if val is not None and isinstance(val, (int, float)):
            entries.append({
                'year': year,
                'month': month,
                'amount': float(val),
                'is_actual': is_actual,
            })
    return entries


def _get_label(ws, row: int, col: int = 1) -> Optional[str]:
    """Get a label from a cell, stripped."""
    val = ws.cell(row=row, column=col).value
    if val is not None:
        return str(val).strip()
    return None


def parse_cashflow_excel(file_path: str) -> dict:
    """
    Parse the cashflow Excel workbook.
    Returns structured data ready for DB insertion:
    {
        'monthly_rows': [
            {id, month, year, category, subcategory, line_item, amount, currency, is_actual, source}
        ],
        'fx_rates': [
            {from_currency, to_currency, rate, year, month}
        ],
        'summary': {
            'total_revenue_rows': int,
            'total_expense_rows': int,
            'total_other_rows': int,
        }
    }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if 'Cashflow' not in wb.sheetnames:
        raise ValueError("Workbook does not contain a 'Cashflow' sheet")

    ws = wb['Cashflow']
    col_dates = _get_column_dates(ws)

    monthly_rows = []
    fx_rates = []

    def add_rows(row_num, category, subcategory=None, line_item_override=None):
        label = line_item_override or _get_label(ws, row_num)
        if not label:
            return
        entries = _read_row_values(ws, row_num, col_dates)
        for entry in entries:
            monthly_rows.append({
                'id': str(uuid.uuid4()),
                'month': entry['month'],
                'year': entry['year'],
                'category': category,
                'subcategory': subcategory,
                'line_item': label,
                'amount': entry['amount'],
                'currency': 'ZAR',
                'is_actual': entry['is_actual'],
                'source': 'excel_import',
            })

    # Summary rows
    add_rows(5, 'Summary', line_item_override='Bank Balance')
    add_rows(6, 'Summary', line_item_override='Net Cash Movement')
    add_rows(7, 'Summary', line_item_override='Operating Profit')
    add_rows(8, 'Summary', line_item_override='Net Income')

    # FX Rate (row 17)
    fx_entries = _read_row_values(ws, 17, col_dates)
    for entry in fx_entries:
        if entry['amount'] > 0:
            fx_rates.append({
                'from_currency': 'ZAR',
                'to_currency': 'NZD',
                'rate': entry['amount'],
                'year': entry['year'],
                'month': entry['month'],
            })

    # Revenue rows (31-42) - individual clients
    for row in range(31, 43):
        label = _get_label(ws, row)
        if label and label.strip():
            add_rows(row, 'Revenue', subcategory='Client Revenue')

    # Total Revenue (row 43)
    add_rows(43, 'Revenue', line_item_override='Total Revenue')

    # Expense rows (48-65)
    for row in range(48, 66):
        label = _get_label(ws, row)
        if label and label.strip():
            add_rows(row, 'Expenses', subcategory='Operating Expenses')

    # Total Expenses (row 66)
    add_rows(66, 'Expenses', line_item_override='Total Expenses')

    # Operating Profit (row 67)
    add_rows(67, 'Summary', line_item_override='Operating Profit (post-expenses)')

    # Earnings Before Tax (row 72)
    add_rows(72, 'Summary', line_item_override='Earnings Before Tax')

    # Pre-Tax Income (row 76)
    add_rows(76, 'Summary', line_item_override='Pre-Tax Income')

    # Loans section (85-97)
    for row in range(85, 98):
        label = _get_label(ws, row)
        if label and label.strip():
            add_rows(row, 'Loans')

    # Vendors section (99-119)
    for row in range(99, 120):
        label = _get_label(ws, row)
        if label and label.strip():
            add_rows(row, 'Vendors')

    # Receivables section (121-127)
    for row in range(121, 128):
        label = _get_label(ws, row)
        if label and label.strip():
            add_rows(row, 'Receivables')

    # Capex (row 131)
    add_rows(131, 'Capex', line_item_override='Capex')

    # Owners section (133-139)
    for row in range(133, 140):
        label = _get_label(ws, row)
        if label and label.strip():
            add_rows(row, 'Owners')

    # Count categories for summary
    revenue_count = sum(1 for r in monthly_rows if r['category'] == 'Revenue')
    expense_count = sum(1 for r in monthly_rows if r['category'] == 'Expenses')
    other_count = len(monthly_rows) - revenue_count - expense_count

    wb.close()

    return {
        'monthly_rows': monthly_rows,
        'fx_rates': fx_rates,
        'summary': {
            'total_revenue_rows': revenue_count,
            'total_expense_rows': expense_count,
            'total_other_rows': other_count,
        }
    }


def parse_daily_cashflow(file_path: str) -> list[dict]:
    """
    Parse the Daily Cashflow sheet from the workbook.
    Returns list of daily entries ready for DB insertion.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if 'Daily Cashflow' not in wb.sheetnames:
        wb.close()
        return []

    ws = wb['Daily Cashflow']
    daily_rows = []

    # Read header row to get column positions
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip().lower()] = col

    # Read data rows starting from row 2
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row, column=headers.get('date', 1)).value
        if date_val is None:
            continue

        # Convert date
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            date_str = str(date_val)

        category = ws.cell(row=row, column=headers.get('category', 2)).value
        line_item = ws.cell(row=row, column=headers.get('line_item', 3)).value
        amount = ws.cell(row=row, column=headers.get('amount', 4)).value
        description = ws.cell(row=row, column=headers.get('description', 5)).value

        if category and amount is not None:
            daily_rows.append({
                'id': str(uuid.uuid4()),
                'date': date_str,
                'category': str(category).strip(),
                'line_item': str(line_item).strip() if line_item else None,
                'amount': float(amount),
                'currency': 'ZAR',
                'description': str(description).strip() if description else None,
                'is_actual': True,
                'source': 'excel_import',
            })

    wb.close()
    return daily_rows
